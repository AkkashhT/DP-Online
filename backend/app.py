from dotenv import load_dotenv
load_dotenv()

import os
import json
import base64
import io
import re
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
from google import genai
from google.genai import types
from PIL import Image
import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import uuid

from database import (
    init_db, save_document, update_document_counts, get_document,
    get_all_documents, delete_document,
    save_extraction, get_extraction,
    update_extraction_text, update_extraction_tables,
    save_image, get_images,
    save_chat_message, get_chat_history, clear_chat_history
)

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'uploads'
RESULTS_FOLDER = 'results'
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULTS_FOLDER, exist_ok=True)

init_db()

GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '')
gemini_client = None
if GEMINI_API_KEY:
    gemini_client = genai.Client(api_key=GEMINI_API_KEY)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def extract_images_from_pdf(pdf_path, doc_id):
    images = []
    img_dir = os.path.join(RESULTS_FOLDER, doc_id, 'images')
    os.makedirs(img_dir, exist_ok=True)
    try:
        doc = fitz.open(pdf_path)
        img_count = 0
        for page_num in range(len(doc)):
            page = doc[page_num]
            for img in page.get_images(full=True):
                xref = img[0]
                base_image = doc.extract_image(xref)
                pil_img = Image.open(io.BytesIO(base_image["image"]))
                if pil_img.mode in ('RGBA', 'P', 'LA'):
                    bg = Image.new('RGB', pil_img.size, (255, 255, 255))
                    if pil_img.mode == 'RGBA':
                        bg.paste(pil_img, mask=pil_img.split()[3])
                    else:
                        bg.paste(pil_img.convert('RGBA'), mask=pil_img.convert('RGBA').split()[3])
                    pil_img = bg
                else:
                    pil_img = pil_img.convert('RGB')
                img_count += 1
                img_filename = f"image_{img_count}.png"
                img_path = os.path.join(img_dir, img_filename)
                pil_img.save(img_path, 'PNG')
                save_image(doc_id, img_count, img_filename, page_num + 1, img_path)
                buf = io.BytesIO()
                pil_img.save(buf, format="PNG")
                images.append({
                    'id': img_count, 'filename': img_filename, 'page': page_num + 1,
                    'data': f"data:image/png;base64,{base64.b64encode(buf.getvalue()).decode()}"
                })
        doc.close()
    except Exception as e:
        print(f"PDF image extraction error: {e}")
    return images


def extract_images_from_image_file(file_path, doc_id):
    images = []
    img_dir = os.path.join(RESULTS_FOLDER, doc_id, 'images')
    os.makedirs(img_dir, exist_ok=True)
    try:
        pil_img = Image.open(file_path).convert('RGB')
        img_path = os.path.join(img_dir, "image_1.png")
        pil_img.save(img_path, 'PNG')
        save_image(doc_id, 1, "image_1.png", 1, img_path)
        buf = io.BytesIO()
        pil_img.save(buf, format="PNG")
        images.append({
            'id': 1, 'filename': "image_1.png", 'page': 1,
            'data': f"data:image/png;base64,{base64.b64encode(buf.getvalue()).decode()}"
        })
    except Exception as e:
        print(f"Image file error: {e}")
    return images


def process_with_gemini(file_path, file_ext):
    if not gemini_client:
        return {'text': 'Gemini API key not configured. Please set GEMINI_API_KEY.', 'tables': []}
    mime_map = {'pdf': 'application/pdf', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png'}
    mime_type = mime_map.get(file_ext, 'application/octet-stream')
    prompt = """Analyze this document and extract all content.

Respond ONLY in this exact JSON format:
{
  "extracted_text": "All text content preserving structure and paragraphs",
  "tables": [
    {
      "title": "Table title or description",
      "headers": ["Column 1", "Column 2"],
      "rows": [["val1", "val2"], ["val3", "val4"]]
    }
  ]
}

Rules:
1. Extract ALL text completely and accurately
2. Find every table — extract headers and all rows
3. If no tables found, return empty array []
4. Return ONLY the JSON — no markdown, no backticks, no extra text"""
    try:
        with open(file_path, 'rb') as f:
            file_bytes = f.read()
        response = gemini_client.models.generate_content(
            model='gemini-2.5-flash',
            contents=[types.Part.from_bytes(data=file_bytes, mime_type=mime_type), prompt]
        )
        text = response.text.strip()
        text = re.sub(r'^```(?:json)?\s*', '', text)
        text = re.sub(r'\s*```$', '', text)
        parsed = json.loads(text)
        return {'text': parsed.get('extracted_text', ''), 'tables': parsed.get('tables', [])}
    except json.JSONDecodeError:
        try:
            return {'text': response.text, 'tables': []}
        except:
            return {'text': 'Could not parse Gemini response.', 'tables': []}
    except Exception as e:
        return {'text': f'Gemini processing error: {str(e)}', 'tables': []}


def chat_with_gemini(question, doc_content, history):
    if not gemini_client:
        return 'Gemini API key not configured.'
    history_text = ""
    if history:
        history_text = "\n\nPREVIOUS CONVERSATION:\n"
        for msg in history[-6:]:
            prefix = "User" if msg['role'] == 'user' else "Assistant"
            history_text += f"{prefix}: {msg['content']}\n"
    prompt = f"""You are a document assistant. Answer based ONLY on the document content below.
If the answer is not in the document, say so clearly.

DOCUMENT CONTENT:
{doc_content}
{history_text}
USER QUESTION: {question}

Answer clearly and accurately based solely on the document."""
    try:
        response = gemini_client.models.generate_content(
            model='gemini-2.5-flash', contents=prompt)
        return response.text
    except Exception as e:
        return f'Chat error: {str(e)}'


@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': 'Unsupported file type. Use PDF, JPG, or PNG'}), 400

    doc_id = str(uuid.uuid4())
    filename = secure_filename(file.filename)
    file_ext = filename.rsplit('.', 1)[1].lower()
    doc_dir = os.path.join(UPLOAD_FOLDER, doc_id)
    os.makedirs(doc_dir, exist_ok=True)
    file_path = os.path.join(doc_dir, filename)
    file.save(file_path)
    file_size = os.path.getsize(file_path)

    save_document(doc_id, filename, file_path, file_ext, file_size)

    images = extract_images_from_pdf(file_path, doc_id) if file_ext == 'pdf' \
        else extract_images_from_image_file(file_path, doc_id)

    result = process_with_gemini(file_path, file_ext)

    doc_content = result['text']
    if result['tables']:
        doc_content += "\n\nTABLES:\n"
        for t in result['tables']:
            doc_content += f"\n{t.get('title', 'Table')}:\n"
            if t.get('headers'):
                doc_content += " | ".join(t['headers']) + "\n"
            for row in t.get('rows', []):
                doc_content += " | ".join(str(c) for c in row) + "\n"

    save_extraction(doc_id, result['text'], result['tables'], doc_content)
    word_count = len(result['text'].split()) if result['text'] else 0
    update_document_counts(doc_id, word_count, len(result['tables']), len(images))

    return jsonify({
        'doc_id': doc_id, 'filename': filename,
        'text': result['text'], 'tables': result['tables'],
        'images': images, 'image_count': len(images), 'table_count': len(result['tables'])
    })


@app.route('/api/edit/text/<doc_id>', methods=['PUT'])
def edit_text(doc_id):
    if not get_extraction(doc_id):
        return jsonify({'error': 'Document not found'}), 404
    data = request.get_json()
    word_count = update_extraction_text(doc_id, data.get('text', ''))
    return jsonify({'success': True, 'word_count': word_count})


@app.route('/api/edit/tables/<doc_id>', methods=['PUT'])
def edit_tables(doc_id):
    if not get_extraction(doc_id):
        return jsonify({'error': 'Document not found'}), 404
    data = request.get_json()
    tables = data.get('tables', [])
    update_extraction_tables(doc_id, tables)
    return jsonify({'success': True, 'table_count': len(tables)})


@app.route('/api/documents', methods=['GET'])
def list_documents():
    return jsonify({'documents': get_all_documents()})


@app.route('/api/documents/<doc_id>', methods=['GET'])
def get_doc(doc_id):
    doc = get_document(doc_id)
    if not doc:
        return jsonify({'error': 'Document not found'}), 404
    extraction = get_extraction(doc_id)
    db_images = get_images(doc_id)
    images_out = []
    for img in db_images:
        entry = {'id': img['image_index'], 'filename': img['filename'],
                 'page': img['page_number'], 'data': None}
        if os.path.exists(img['file_path']):
            with open(img['file_path'], 'rb') as f:
                entry['data'] = f"data:image/png;base64,{base64.b64encode(f.read()).decode()}"
        images_out.append(entry)
    return jsonify({
        'doc_id': doc_id, 'filename': doc['filename'],
        'uploaded_at': doc['uploaded_at'], 'word_count': doc['word_count'],
        'table_count': doc['table_count'], 'image_count': doc['image_count'],
        'text': extraction['text'] if extraction else '',
        'tables': extraction['tables'] if extraction else [],
        'images': images_out,
    })


@app.route('/api/documents/<doc_id>', methods=['DELETE'])
def delete_doc(doc_id):
    if not get_document(doc_id):
        return jsonify({'error': 'Document not found'}), 404
    delete_document(doc_id)
    return jsonify({'success': True})


@app.route('/api/results/<doc_id>', methods=['GET'])
def get_results(doc_id):
    return get_doc(doc_id)


@app.route('/api/download/text/<doc_id>', methods=['GET'])
def download_text(doc_id):
    extraction = get_extraction(doc_id)
    if not extraction:
        return jsonify({'error': 'Document not found'}), 404
    buf = io.BytesIO()
    buf.write((extraction['text'] or '').encode('utf-8'))
    buf.seek(0)
    return send_file(buf, mimetype='text/plain', as_attachment=True, download_name='extracted_text.txt')


@app.route('/api/download/tables/<doc_id>', methods=['GET'])
def download_tables(doc_id):
    extraction = get_extraction(doc_id)
    if not extraction:
        return jsonify({'error': 'Document not found'}), 404
    tables = extraction['tables']
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if not tables:
        ws = wb.create_sheet("No Tables Found")
        ws['A1'] = "No tables were found in this document."
    else:
        for t_idx, table in enumerate(tables):
            ws = wb.create_sheet(table.get('title', f'Table {t_idx+1}')[:31])
            hfill = PatternFill("solid", fgColor="2563EB")
            hfont = Font(bold=True, color="FFFFFF", size=11)
            thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            row_offset = 1
            if table.get('title'):
                ws.cell(1, 1, table['title']).font = Font(bold=True, size=13)
                row_offset = 2
            for ci, h in enumerate(table.get('headers', []), 1):
                c = ws.cell(row_offset, ci, h)
                c.fill = hfill; c.font = hfont
                c.alignment = Alignment(horizontal='center'); c.border = thin
            for ri, row in enumerate(table.get('rows', []), row_offset + 1):
                for ci, val in enumerate(row, 1):
                    c = ws.cell(ri, ci, str(val))
                    c.border = thin
                    if ri % 2 == 0:
                        c.fill = PatternFill("solid", fgColor="EFF6FF")
            for col in ws.columns:
                w = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(w + 4, 40)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='tables.xlsx')


@app.route('/api/download/image/<doc_id>/<int:image_id>', methods=['GET'])
def download_image(doc_id, image_id):
    db_images = get_images(doc_id)
    target = next((i for i in db_images if i['image_index'] == image_id), None)
    if not target or not os.path.exists(target['file_path']):
        return jsonify({'error': 'Image not found'}), 404
    return send_file(target['file_path'], mimetype='image/png',
                     as_attachment=True, download_name=target['filename'])


@app.route('/api/chat/<doc_id>', methods=['POST'])
def chat(doc_id):
    extraction = get_extraction(doc_id)
    if not extraction:
        return jsonify({'error': 'Document not found'}), 404
    data = request.get_json()
    question = data.get('question', '').strip()
    if not question:
        return jsonify({'error': 'No question provided'}), 400
    history = get_chat_history(doc_id)
    answer = chat_with_gemini(question, extraction['doc_content'], history)
    save_chat_message(doc_id, 'user', question)
    save_chat_message(doc_id, 'assistant', answer)
    return jsonify({'answer': answer})


@app.route('/api/chat/<doc_id>/history', methods=['GET'])
def chat_history(doc_id):
    return jsonify({'history': get_chat_history(doc_id)})


@app.route('/api/chat/<doc_id>/clear', methods=['DELETE'])
def clear_chat(doc_id):
    clear_chat_history(doc_id)
    return jsonify({'success': True})


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'gemini_configured': bool(GEMINI_API_KEY)})


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
